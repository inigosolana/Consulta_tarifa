"""
knowledge_feedback.py
=====================
Módulo de "Bucle de Retroalimentación de Conocimiento" para el agente de voz.

Responsabilidades:
  1. Tool de búsqueda de tarifas en el Excel (pandas).
  2. Extracción de información comercial de la transcripción via LLM.
  3. Actualización del Excel con el nuevo conocimiento (preservando todas las hojas).
"""

from __future__ import annotations

import logging
import os
from typing import Annotated

import pandas as pd
from openpyxl import load_workbook
from openai import AsyncOpenAI

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Configuración
# ---------------------------------------------------------------------------

EXCEL_PATH = os.path.join(
    os.path.dirname(__file__),
    "informe_servicios 20260313.xlsx",
)
SHEET_NAME = "informe CRM"
COL_NOMBRE = "Nombre"
COL_INFO   = "Información comercial"

_openai_client: AsyncOpenAI | None = None


def _get_openai_client() -> AsyncOpenAI:
    global _openai_client
    if _openai_client is None:
        _openai_client = AsyncOpenAI()
    return _openai_client


# ---------------------------------------------------------------------------
# Helpers de Excel
# ---------------------------------------------------------------------------

def _load_df() -> pd.DataFrame:
    """Carga la hoja 'informe CRM' del Excel."""
    return pd.read_excel(EXCEL_PATH, sheet_name=SHEET_NAME, dtype=str, engine="openpyxl")


def _save_cell(nombre_tarifa: str, nuevo_valor: str) -> bool:
    """
    Actualiza SOLO la celda de 'Información comercial' de la tarifa indicada,
    preservando el resto de hojas y el formato del Excel sin tocarlo.
    Devuelve True si se encontró y actualizó la fila.
    """
    wb = load_workbook(EXCEL_PATH)
    ws = wb[SHEET_NAME]

    # Localizar la columna Nombre y la columna Información comercial
    headers = {cell.value: cell.column for cell in ws[1]}
    col_nombre = headers.get(COL_NOMBRE)
    col_info   = headers.get(COL_INFO)

    if col_nombre is None or col_info is None:
        logger.error(
            "No se encontraron las columnas '%s' o '%s' en la hoja '%s'.",
            COL_NOMBRE, COL_INFO, SHEET_NAME,
        )
        wb.close()
        return False

    # Buscar la fila (ignorar mayúsculas/minúsculas y espacios)
    nombre_buscado = nombre_tarifa.strip().lower()
    fila_encontrada: int | None = None

    for row in ws.iter_rows(min_row=2, values_only=False):
        celda_nombre = row[col_nombre - 1]
        if celda_nombre.value and str(celda_nombre.value).strip().lower() == nombre_buscado:
            fila_encontrada = celda_nombre.row
            break

    if fila_encontrada is None:
        logger.warning("Fila para '%s' no encontrada al intentar guardar.", nombre_tarifa)
        wb.close()
        return False

    ws.cell(row=fila_encontrada, column=col_info).value = nuevo_valor
    wb.save(EXCEL_PATH)
    wb.close()
    return True


# ---------------------------------------------------------------------------
# 1. TOOL: buscar_tarifa
# ---------------------------------------------------------------------------

async def buscar_tarifa(
    nombre_tarifa: Annotated[
        str,
        "Nombre exacto o aproximado de la tarifa que el comercial ha preguntado.",
    ],
) -> str:
    """
    Busca una tarifa en el CRM por su nombre.

    Devuelve la información comercial si existe, o una instrucción para
    transferir la llamada si el campo está vacío.
    """
    try:
        df = _load_df()
    except FileNotFoundError:
        logger.error("Excel no encontrado en: %s", EXCEL_PATH)
        return (
            "No puedo acceder al CRM en este momento. "
            "Por favor, transfiere la llamada a un agente humano."
        )

    nombre_limpio = nombre_tarifa.strip().lower()

    # 1º intento: coincidencia exacta
    mask = df[COL_NOMBRE].str.strip().str.lower() == nombre_limpio
    fila = df[mask]

    # 2º intento: coincidencia parcial (contiene)
    if fila.empty:
        mask = df[COL_NOMBRE].str.lower().str.contains(nombre_limpio, na=False)
        fila = df[mask]

    if fila.empty:
        logger.info("Tarifa no encontrada en CRM: '%s'", nombre_tarifa)
        return (
            f"No encontré la tarifa '{nombre_tarifa}' en el CRM. "
            "Voy a transferirte con un agente humano que podrá ayudarte."
        )

    info = fila.iloc[0][COL_INFO]

    # Campo vacío, NaN o solo espacios → transferir
    if pd.isna(info) or str(info).strip() in ("", "nan"):
        nombre_real = fila.iloc[0][COL_NOMBRE]
        logger.info(
            "Tarifa '%s' encontrada pero sin información comercial.", nombre_real
        )
        return (
            f"Tengo registrada la tarifa '{nombre_real}' pero aún no "
            "dispongo de su información comercial detallada. "
            "Voy a transferirte con un agente humano para que te lo explique."
        )

    nombre_real = fila.iloc[0][COL_NOMBRE]
    logger.info("Tarifa '%s' encontrada. Devolviendo información al agente.", nombre_real)
    return str(info).strip()


# ---------------------------------------------------------------------------
# 2. Extractor LLM
# ---------------------------------------------------------------------------

_EXTRACTION_PROMPT = """\
Eres un asistente especializado en extraer información comercial de transcripciones de llamadas.

Se te proporcionará la transcripción completa de una llamada en la que un agente humano \
explicó a un comercial los detalles de una tarifa telefónica específica.

Tu única tarea es extraer, de forma concisa y fiel, la explicación técnico-comercial que \
el agente humano dio sobre la tarifa. Incluye precios, condiciones, velocidades, \
permanencias y cualquier dato relevante que el agente haya mencionado.

REGLAS ESTRICTAS:
- Si la transcripción NO contiene una explicación clara de la tarifa (llamada cortada, \
  tema diferente, conversación incompleta o sin datos útiles), responde ÚNICAMENTE con: N/A
- No añadas introducciones, despedidas ni comentarios propios.
- Responde en español.
- Máximo 400 palabras.

TARIFA CONSULTADA: {nombre_tarifa}

TRANSCRIPCIÓN:
{transcripcion}
"""


async def analizar_transcripcion(nombre_tarifa: str, transcripcion: str) -> str:
    """
    Usa GPT para extraer la información comercial de la transcripción.
    Devuelve el texto extraído o "N/A" si no hay información válida.
    """
    prompt = _EXTRACTION_PROMPT.format(
        nombre_tarifa=nombre_tarifa,
        transcripcion=transcripcion,
    )

    client = _get_openai_client()
    logger.info(
        "Enviando transcripción al LLM (%d chars) para tarifa '%s'...",
        len(transcripcion),
        nombre_tarifa,
    )

    response = await client.chat.completions.create(
        model="gpt-4o-mini",
        messages=[{"role": "user", "content": prompt}],
        temperature=0.2,
        max_tokens=600,
    )

    resultado = response.choices[0].message.content.strip()
    logger.info("LLM extrajo: %.120s...", resultado)
    return resultado


# ---------------------------------------------------------------------------
# 3. Actualizar Excel con nuevo conocimiento
# ---------------------------------------------------------------------------

async def actualizar_crm(nombre_tarifa: str, nueva_info: str) -> bool:
    """
    Escribe `nueva_info` en la celda 'Información comercial' de `nombre_tarifa`
    en el Excel, preservando todas las demás hojas y el formato.

    Devuelve True si la actualización fue exitosa.
    """
    if not nueva_info or nueva_info.strip().upper() == "N/A":
        logger.info(
            "LLM no extrajo información válida para '%s'. Excel no modificado.",
            nombre_tarifa,
        )
        return False

    actualizado = _save_cell(nombre_tarifa, nueva_info)

    if actualizado:
        logger.info(
            "=== CRM ACTUALIZADO === Tarifa: '%s' | %d caracteres guardados.",
            nombre_tarifa,
            len(nueva_info),
        )
    else:
        logger.warning(
            "No se pudo actualizar el CRM para '%s'.", nombre_tarifa
        )

    return actualizado


# ---------------------------------------------------------------------------
# 4. Función orquestadora post-llamada
# ---------------------------------------------------------------------------

async def procesar_post_llamada(nombre_tarifa: str, transcripcion: str) -> None:
    """
    Punto de entrada para el evento de desconexión de la sala.

    1. Envía la transcripción al LLM para extraer la información comercial.
    2. Si el LLM devuelve contenido válido, actualiza el Excel.
    """
    logger.info(
        "--- POST-LLAMADA: procesando transcripción para tarifa '%s' ---",
        nombre_tarifa,
    )

    if not transcripcion.strip():
        logger.info("Transcripción vacía. Nada que procesar.")
        return

    nueva_info = await analizar_transcripcion(nombre_tarifa, transcripcion)
    actualizado = await actualizar_crm(nombre_tarifa, nueva_info)

    if actualizado:
        print(
            f"\n✔ [CRM] Tarifa '{nombre_tarifa}' actualizada con nueva "
            f"información comercial ({len(nueva_info)} chars).\n"
        )
    else:
        print(
            f"\n✘ [CRM] No se actualizó el CRM para '{nombre_tarifa}' "
            f"(sin información válida o tarifa no encontrada).\n"
        )
